"""
XLSX 电子表格解析器 — 单一职责：将 Excel 解析为增强文本（sheet 标题 + HTML 表格）。

使用 openpyxl 读取 .xlsx / .xls 文件：
    - 每个 sheet 输出为 <h2>sheet 名</h2> 标题 + HTML <table>；
    - 第一行视为表头（<th>），其余为数据行（<td>）；
    - 空行自动跳过，纯空白 sheet 只输出标题；
    - 列宽对齐：合并单元格导致行长度不一致时补空列。

降级旁路（对齐图片流程的 officeparser 兜底）：
    - openpyxl 解析失败 → 降级到 pandas.read_excel（支持 .xls / 损坏文件）；
    - pandas 也失败 → 返回空字符串，调用方降级为纯文本提取。

遵循优雅降级：
    - openpyxl 未安装 → 直接尝试 pandas；
    - 单个 sheet 解析异常 → 跳过该 sheet，继续处理其余 sheet；
    - 配置开关可控制表格提取、行数和 sheet 数量上限。
"""

from __future__ import annotations

from typing import Any

from app.config import get_settings
from app.document.base import DocumentParser, ParsedSection
from app.utils.logger import get_logger

log = get_logger(__name__)


class XLSXParser(DocumentParser):
    """XLSX 解析器 — openpyxl 优先 + pandas 降级，每 sheet 转 HTML 表格。"""

    async def parse(self, file_path: str) -> str:
        """解析 XLSX 文档，返回增强文本。

        解析流程：
            1. 尝试 openpyxl（read_only + data_only）；
            2. openpyxl 失败 → 降级到 pandas.read_excel；
            3. 两者都失败 → 返回空字符串。

        Args:
            file_path: XLSX 文件路径。

        Returns:
            增强文本（<h2>sheet 名</h2> + HTML 表格）。
            所有解析器都失败时返回空字符串。
        """
        settings = get_settings()
        table_enabled = self._bool(
            getattr(settings, "XLSX_TABLE_EXTRACTION_ENABLED", True), True
        )
        max_rows = self._int(getattr(settings, "XLSX_MAX_ROWS_PER_SHEET", 500), 500)
        max_sheets = self._int(getattr(settings, "XLSX_MAX_SHEETS", 20), 20)

        # 优先尝试 openpyxl
        sections = self._parse_with_openpyxl(
            file_path, table_enabled, max_rows, max_sheets
        )

        # openpyxl 失败时降级到 pandas
        if sections is None:
            log.info("xlsx.fallback_to_pandas", file_path=file_path)
            sections = self._parse_with_pandas(
                file_path, table_enabled, max_rows, max_sheets
            )

        if sections is None:
            # 两个引擎都失败
            return ""

        log.info(
            "xlsx.parsed",
            file_path=file_path,
            sheets=sum(1 for s in sections if s.kind == "text" and "<h2>" in s.content),
            sections=len(sections),
            tables=sum(1 for s in sections if s.kind == "table"),
        )
        return self.sections_to_text(sections)

    def _parse_with_openpyxl(
        self,
        file_path: str,
        table_enabled: bool,
        max_rows: int,
        max_sheets: int,
    ) -> list[ParsedSection] | None:
        """使用 openpyxl 解析 XLSX。

        Returns:
            ParsedSection 列表。openpyxl 未安装或打开失败时返回 None（触发降级）。
        """
        try:
            import openpyxl
        except ImportError:
            log.warning("xlsx.openpyxl_not_installed")
            return None

        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as exc:
            log.warning("xlsx.openpyxl_open_failed", file_path=file_path, error=str(exc))
            return None

        sections: list[ParsedSection] = []
        sheet_count = 0

        try:
            for sheet_name in workbook.sheetnames:
                if sheet_count >= max_sheets:
                    log.info("xlsx.max_sheets_reached", limit=max_sheets)
                    break

                sheet_count += 1
                seq = sheet_count

                # 输出 sheet 标题（作为分块锚点）
                sections.append(
                    ParsedSection(
                        kind="text",
                        content=f"<h2>{self._escape_html(sheet_name)}</h2>",
                        page=seq,
                    )
                )

                if not table_enabled:
                    continue

                # 提取 sheet 数据为 HTML 表格
                try:
                    worksheet = workbook[sheet_name]
                    html = self._extract_sheet_html(worksheet, max_rows)
                    if html:
                        sections.append(
                            ParsedSection(kind="table", content=html, page=seq)
                        )
                except Exception as exc:
                    log.debug(
                        "xlsx.sheet_extract_failed",
                        sheet=sheet_name,
                        error=str(exc),
                    )
        finally:
            workbook.close()

        return sections

    def _parse_with_pandas(
        self,
        file_path: str,
        table_enabled: bool,
        max_rows: int,
        max_sheets: int,
    ) -> list[ParsedSection] | None:
        """pandas 降级解析 — openpyxl 失败时的兜底。

        pandas.read_excel 支持 .xls 格式和部分损坏的 .xlsx 文件，
        使用 openpyxl 或 xlrd 作为后端引擎。

        Returns:
            ParsedSection 列表。pandas 未安装或解析失败时返回 None。
        """
        try:
            import pandas as pd
        except ImportError:
            log.warning("xlsx.pandas_not_installed")
            return None

        try:
            # sheet_name=None 返回所有 sheet 的 dict {sheet_name: DataFrame}
            all_sheets = pd.read_excel(file_path, sheet_name=None)
        except Exception as exc:
            log.warning("xlsx.pandas_open_failed", file_path=file_path, error=str(exc))
            return None

        sections: list[ParsedSection] = []
        sheet_count = 0

        for sheet_name, df in all_sheets.items():
            if sheet_count >= max_sheets:
                log.info("xlsx.max_sheets_reached", limit=max_sheets)
                break

            sheet_count += 1
            seq = sheet_count

            # 输出 sheet 标题
            sections.append(
                ParsedSection(
                    kind="text",
                    content=f"<h2>{self._escape_html(str(sheet_name))}</h2>",
                    page=seq,
                )
            )

            if not table_enabled:
                continue

            # DataFrame → HTML 表格
            try:
                # 限制行数
                df_limited = df.head(max_rows)
                # 跳过空 DataFrame
                if df_limited.empty:
                    continue

                # 转为行列表（含列名作为表头）
                rows: list[list[str | None]] = []
                # 表头行
                rows.append([str(col) for col in df_limited.columns])
                # 数据行
                for _, row in df_limited.iterrows():
                    cells = [
                        str(v).strip() if pd.notna(v) else ""
                        for v in row.tolist()
                    ]
                    # 跳过全空行
                    if all(c == "" for c in cells):
                        continue
                    rows.append(cells)

                if rows:
                    html = self._rows_to_html(rows)
                    if html:
                        sections.append(
                            ParsedSection(kind="table", content=html, page=seq)
                        )
            except Exception as exc:
                log.debug(
                    "xlsx.pandas_sheet_failed",
                    sheet=sheet_name,
                    error=str(exc),
                )

        return sections

    def _extract_sheet_html(self, worksheet: Any, max_rows: int) -> str:
        """将 worksheet 转为 HTML <table>。

        Args:
            worksheet: openpyxl Worksheet 对象。
            max_rows: 最大提取行数。

        Returns:
            HTML <table> 标签字符串。空 sheet 返回空字符串。
        """
        rows: list[list[str | None]] = []
        max_cols = 0
        row_count = 0

        for row in worksheet.iter_rows(values_only=True):
            if row_count >= max_rows:
                log.info("xlsx.max_rows_reached", limit=max_rows)
                break
            row_count += 1

            # 跳过全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # 将所有单元格转为字符串
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            # 跟踪最大列数（合并单元格场景行长度可能不一致）
            if len(cells) > max_cols:
                max_cols = len(cells)
            rows.append(cells)

        if not rows:
            return ""

        # 列宽对齐 — 补齐短行为最大列数（处理合并单元格）
        rows = [r + [""] * (max_cols - len(r)) for r in rows]

        return self._rows_to_html(rows)

    @staticmethod
    def _rows_to_html(rows: list[list[str | None]]) -> str:
        """将二维数据转为 HTML <table> 标签。

        第一行视为表头（<th>），其余为数据行（<td>）。
        自动补齐不等长行（合并单元格场景）。
        """
        if not rows:
            return ""

        # 找出最大列数并补齐
        max_cols = max(len(r) for r in rows) if rows else 0
        padded_rows = [list(r) + [""] * (max_cols - len(r)) for r in rows]

        lines: list[str] = ["<table>"]

        for i, row in enumerate(padded_rows):
            lines.append("<tr>")
            tag = "th" if i == 0 else "td"
            for cell in row:
                cell_text = (cell or "").strip()
                cell_text = XLSXParser._escape_html(cell_text)
                lines.append(f"<{tag}>{cell_text}</{tag}>")
            lines.append("</tr>")

        lines.append("</table>")
        return "\n".join(lines)

    @staticmethod
    def _escape_html(text: str) -> str:
        """转义 HTML 特殊字符。"""
        if not text:
            return ""
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
